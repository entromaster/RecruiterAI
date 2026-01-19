import zipfile
import pandas as pd
from io import StringIO
import os
import openpyxl
import re
import json

import re
from datetime import datetime

# Map custom placeholders to Python's datetime directives
PLACEHOLDER_MAP = [
    ("MMMM", "%B"),   # full month name (January)
    ("MMM", "%b"),    # abbreviated month name (Jan)
    ("YYYY", "%Y"),   # four-digit year (2023)
    ("YY", "%y"),     # two-digit year (23)
    ("MM", "%m"),     # two-digit month (01-12)
    ("DD", "%d"),     # two-digit day (01-31)
]

class DateFormatNormalizer:
    """
    A class to handle date format normalization between different formats based on configuration.
    """

    DATE_COMPONENTS = {
        'MM': 2,
        'DD': 2,
        'YYYY': 4,
        'YY': 2
    }

    def __init__(self, config):
        self.config = config
        self.footer_components = self._parse_format(config['footer']['format'])
        self.table_components = self._parse_format(config['table']['format'])

    def _parse_format(self, format_str):
        components = []
        current = ''
        for char in format_str:
            if char == '/':
                if current:
                    components.append(current)
                    current = ''
            else:
                current += char
        if current:
            components.append(current)
        return components

    def _validate_format(self, date_str, components):
        """Validate that the input string matches the expected format exactly"""
        import re
        # Build regex pattern based on components
        pattern_map = {
            'MM': r'\d{2}',
            'DD': r'\d{2}',
            'YYYY': r'\d{4}',
            'YY': r'\d{2}'
        }
        pattern = '^' + '/'.join(pattern_map[comp] for comp in components) + '$'
        return bool(re.match(pattern, date_str))

    def _convert_date(self, date_str, from_format, to_format):
        try:
            date_obj = datetime.strptime(date_str, from_format)
            return date_obj.strftime(to_format)
        except ValueError:
            return None

    def _format_to_strftime(self, components):
        format_map = {
            'MM': '%m',
            'DD': '%d',
            'YYYY': '%Y',
            'YY': '%y'
        }
        return '/'.join(format_map[comp] for comp in components)

    def _should_normalize(self):
        footer_precision = sum(self.DATE_COMPONENTS[comp] for comp in self.footer_components)
        table_precision = sum(self.DATE_COMPONENTS[comp] for comp in self.table_components)
        return footer_precision < table_precision

    def normalize_table_date(self, date_str):
        # First validate the input format
        if not self._validate_format(date_str, self.table_components):
            return None

        if not self._should_normalize():
            return date_str

        from_format = self._format_to_strftime(self.table_components)
        to_format = self._format_to_strftime(self.footer_components)

        return self._convert_date(date_str, from_format, to_format)

def process_dates_in_table(form_number_table: pd.DataFrame, config: dict) -> pd.DataFrame:
    """
    Process dates in the form table, handling both separate date columns and embedded dates.

    Parameters:
    form_number_table (pd.DataFrame): The table containing form numbers and potentially dates
    config (dict): Configuration dictionary with date settings

    Returns:
    pd.DataFrame: Processed DataFrame with standardized dates
    """
    try:
        if form_number_table is None or form_number_table.empty:
            return form_number_table

        date_config = config.get("DateSubConfig")
        if not date_config:
            return form_number_table

        # Identify the date handling strategy
        date_column, is_embedded = identify_date_column_strategy(form_number_table, config)

        if not date_column:
            print("No date column or embedded dates found")
            return form_number_table

        # Create date normalizer instance
        try:
            print("Creating normalizer with config:", date_config)
            normalizer = DateFormatNormalizer(date_config)
            print("Normalizer created successfully")
        except Exception as e:
            print(f"Failed to create normalizer: {str(e)}")
            return form_number_table

        if is_embedded:
            print(f"Processing embedded dates in column: {date_column}")
            return process_embedded_dates(form_number_table, date_column, date_config, normalizer)
        else:
            print(f"Processing separate date column: {date_column}")
            return process_separate_date_column(form_number_table, date_column, date_config, normalizer)

    except Exception as e:
        print(f"Error in process_dates_in_table: {str(e)}")
        return form_number_table

def process_separate_date_column(form_number_table: pd.DataFrame, date_column: str,
                               date_config: dict, normalizer: DateFormatNormalizer) -> pd.DataFrame:
    """Process dates in a separate date column using the normalizer"""
    try:
        result_df = form_number_table.copy()

        # Process each date in the column
        for idx, row in result_df.iterrows():
            date_str = str(row[date_column]).strip()
            if date_str:
                normalized_date = normalizer.normalize_table_date(date_str)
                if normalized_date:
                    result_df.at[idx, date_column] = normalized_date

        return result_df

    except Exception as e:
        print(f"Error processing separate date column: {str(e)}")
        return form_number_table

def process_embedded_dates(form_number_table: pd.DataFrame, form_number_col: str,
                         date_config: dict, normalizer: DateFormatNormalizer) -> pd.DataFrame:
    """
    Process form numbers that have embedded dates using the normalizer.
    """
    try:
        result_df = form_number_table.copy()

        if 'EDITION DATE' not in result_df.columns:
            result_df['EDITION DATE'] = None

        # Process each form number
        for idx, row in result_df.iterrows():
            form_num = str(row[form_number_col]).strip()
            clean_form_num, date_str, _ = extract_date_from_form_number(form_num, date_config)

            if date_str:
                normalized_date = normalizer.normalize_table_date(date_str)
                if normalized_date:
                    result_df.at[idx, form_number_col] = clean_form_num
                    result_df.at[idx, 'EDITION DATE'] = normalized_date

        return result_df

    except Exception as e:
        print(f"Error processing embedded dates: {str(e)}")
        return form_number_table

def convert_format_to_strptime(fmt: str) -> str:
    """Convert our custom date format to strptime format"""
    replacements = {
        'MM': '%m',
        'DD': '%d',
        'YYYY': '%Y',
        'YY': '%y'
    }
    for custom, strp in replacements.items():
        fmt = fmt.replace(custom, strp)
    return fmt

def extract_date_from_form_number(form_number: str, date_config: dict) -> tuple[str, str, str]:
    """
    Extract date from form number string when no separate date column exists.
    Returns (form_number_without_date, date_str, standardized_date) or (original_form_number, None, None) if no date found.
    """
    try:
        print(f"\nDEBUG: extract_date_from_form_number called with:")
        print(f"form_number: {form_number}")
        print(f"date_config: {json.dumps(date_config, indent=2)}")

        if not form_number or not date_config:
            print("DEBUG: Empty form number or config")
            return form_number, None, None

        # Get patterns and formats
        table_pattern = date_config.get("table", {}).get("pattern")
        table_format = date_config.get("table", {}).get("format")
        footer_format = date_config.get("footer", {}).get("format")

        print(f"DEBUG: Using table_pattern: {table_pattern}")
        print(f"DEBUG: Using table_format: {table_format}")
        print(f"DEBUG: Using footer_format: {footer_format}")

        if not table_pattern or not table_format or not footer_format:
            print("DEBUG: Missing pattern or format")
            return form_number, None, None

        # Convert formats to strptime format
        table_strptime = convert_format_to_strptime(table_format)
        footer_strptime = convert_format_to_strptime(footer_format)

        print(f"DEBUG: Converted formats - table: {table_strptime}, footer: {footer_strptime}")

        # Find the last opening parenthesis before the date
        paren_idx = form_number.rfind('(')
        if paren_idx == -1:
            print("DEBUG: No parentheses found")
            return form_number, None, None

        # Extract date using the table pattern
        match = re.search(table_pattern, form_number)
        print(f"DEBUG: Regex match result: {match.groups() if match else None}")

        if not match:
            print("DEBUG: No date match found")
            return form_number, None, None

        date_str = match.group(1)
        print(f"DEBUG: Extracted date_str: {date_str}")

        # Get the form number without the date and parentheses
        form_number_clean = form_number[:paren_idx].strip()
        print(f"DEBUG: Cleaned form number: {form_number_clean}")

        # Parse using table format first, then convert to footer format
        try:
            # Parse with table format
            date_obj = datetime.strptime(date_str, table_strptime)
            print(f"DEBUG: Parsed date object: {date_obj}")

            # Format to footer format
            standardized_date = datetime.strftime(date_obj, footer_strptime)
            print(f"DEBUG: Standardized date: {standardized_date}")

            return form_number_clean, date_str, standardized_date
        except Exception as e:
            print(f"DEBUG: Error standardizing date: {str(e)}")
            return form_number, None, None

    except Exception as e:
        print(f"DEBUG: Error in extract_date_from_form_number: {str(e)}")
        return form_number, None, None



def identify_date_column_strategy(form_number_table: pd.DataFrame, config: dict) -> tuple[str, bool]:
    """
    Identify whether dates are in a separate column or embedded in form numbers.

    Parameters:
    form_number_table (pd.DataFrame): The table containing form numbers
    config (dict): Configuration dictionary with date settings

    Returns:
    tuple: (column_name, is_embedded) where is_embedded indicates if dates are in form numbers
    """
    try:
        # Get column aliases from config
        column_requirements = config.get("TableConfig", {}).get("ColumnRequirements", {})
        edition_date_aliases = column_requirements.get("EDITION DATE", ["EDITION DATE"])

        # Use table pattern for checking embedded dates - this is the input format
        table_pattern = config.get("DateSubConfig", {}).get("table", {}).get("pattern")

        print(f"DEBUG identify_strategy - Table pattern: {table_pattern}")

        # Check for separate date column first
        for date_alias in edition_date_aliases:
            if date_alias in form_number_table.columns:
                # Verify the column contains valid dates
                sample_dates = form_number_table[date_alias].dropna().astype(str).head()
                if table_pattern:
                    for date_str in sample_dates:
                        if re.search(table_pattern, str(date_str).strip()):
                            print(f"DEBUG identify_strategy - Found separate date column: {date_alias}")
                            return date_alias, False

        # If no separate date column, check form number column for embedded dates
        form_number_col = None
        for alias in column_requirements.get("FORM NUMBER", ["FORM NUMBER"]):
            if alias in form_number_table.columns:
                form_number_col = alias
                break

        if form_number_col and table_pattern:
            # Check if any form numbers contain dates in the input format
            sample_forms = form_number_table[form_number_col].dropna().astype(str).head()
            for form_num in sample_forms:
                if re.search(table_pattern, str(form_num).strip()):
                    print(f"DEBUG identify_strategy - Found embedded dates in column: {form_number_col}")
                    return form_number_col, True

        print("DEBUG identify_strategy - No date column or embedded dates found")
        return None, False

    except Exception as e:
        print(f"Error identifying date column strategy: {str(e)}")
        return None, False

def convert_custom_format_to_strptime(fmt: str) -> str:
    """
    Convert a custom date format (with placeholders like MM, DD, YYYY, MMM, MMMM, YY)
    into a Python datetime-compatible format for `strptime`/`strftime`.
    """
    python_fmt = fmt
    for placeholder, directive in PLACEHOLDER_MAP:
        if placeholder in python_fmt:
            python_fmt = python_fmt.replace(placeholder, directive)
    return python_fmt

def parse_date_with_possible_defaults(date_str: str, source_format: str) -> datetime:
    """
    Parse a date string using the expanded placeholder map.
    - If the format omits day (e.g., "MM/YYYY" or "MMMM YYYY"), default day to 1.
    - If the format uses two-digit year (YY), Python's strptime will handle it with century pivot.
    - If the format omits month or year, you can decide how to default them (currently year=1900 if missing).
    """
    python_fmt = convert_custom_format_to_strptime(source_format)

    # Check if "DD" is in the original source_format
    has_day = "DD" in source_format

    if not has_day:
        python_fmt = python_fmt.replace("%d", "")
        # Clean up separators
        python_fmt = re.sub(r'[/\-\.]+$', '', python_fmt)  # trailing
        python_fmt = re.sub(r'([/\-\.])\1+', r'\1', python_fmt)  # repeated

    date_obj = datetime.strptime(date_str.strip(), python_fmt.strip())

    if not has_day:
        date_obj = date_obj.replace(day=1)

    return date_obj

def format_date_with_custom_placeholders(date_obj: datetime, target_format: str) -> str:
    """
    Convert a `datetime` object to a string using placeholders like MM, DD, YYYY, YY, MMM, MMMM.
    """
    python_fmt = convert_custom_format_to_strptime(target_format)
    return date_obj.strftime(python_fmt)

def extract_and_standardize_date(text: str, pattern: str, source_format: str, target_format: str) -> tuple[str, str]:
    """
    Extract date from text using `pattern` and convert it from `source_format` to `target_format`.
    Returns (original_match, standardized_date) or (None, None) if no match.
    """
    try:
        match = re.search(pattern, text)
        if not match:
            return None, None

        date_str = match.group(1)  # The date portion from the regex group

        # Parse the date string with our improved parser
        date_obj = parse_date_with_possible_defaults(date_str, source_format)

        # Format using the target format
        standardized = format_date_with_custom_placeholders(date_obj, target_format)

        return match.group(0), standardized
    except Exception as e:
        print(f"Error standardizing date: {str(e)}")
        return None, None

def standardize_form_number_with_date(form_number: str, date_config: dict) -> str:
    """
    Standardize a form number that includes a date based on the date configuration.
    """
    try:
        if not date_config:
            return form_number

        original_date, standardized_date = extract_and_standardize_date(
            form_number,
            date_config["footer"]["pattern"],
            date_config["footer"]["format"],
            date_config["table"]["format"]
        )

        if original_date and standardized_date:
            return form_number.replace(original_date, standardized_date)

        return form_number
    except Exception as e:
        print(f"Error in form number standardization: {str(e)}")
        return form_number

def combine_form_number_with_date(form_number: str, date_str: str, date_config: dict) -> str:
    """
    Combine form number with a date string according to the configuration.
    The date string should already be in the correct format (including parentheses if needed).

    Parameters:
    form_number (str): The form number to combine
    date_str (str): The standardized date string
    date_config (dict): Date configuration dictionary

    Returns:
    str: Combined form number and date
    """
    try:
        if not date_str or not date_config:
            return form_number

        # Clean up any extra spaces
        clean_form_number = form_number.strip()
        clean_date = date_str.strip()

        # Combine with a single space
        return f"{clean_form_number} {clean_date}"
    except Exception as e:
        print(f"Error combining form number with date: {str(e)}")
        return form_number


# Text Mode Form Number Processing Code
def preprocess_text(text):
    # Preprocess the text by inserting spaces around "and" if it concatenates form numbers
    text = re.sub(r"(\(\d{2}/\d{2}\))and(\w)", r"\1 and \2", text)
    return text


def extract_text_from_adobe_json(data):
    text_list = []

    def extract_text(item):
        # If 'Text' exists in the current level, add it to the list
        if "Text" in item:
            text_list.append(item["Text"])
        # If 'Kids' exists, recursively call extract_text on each child
        if "Kids" in item:
            for kid in item["Kids"]:
                extract_text(kid)

    # Process each top-level item in the data list
    for entry in data:
        extract_text(entry)

    return text_list


def filter_adobe_json_by_object_id(adobe_json, start_index=None, end_index=None):
    # Set start_index to minimum ObjectID if None
    if start_index is None:
        start_index = 0

    # Set end_index to maximum ObjectID if None
    if end_index is None:
        end_index = -1

    # Filter adobe_json list based on the adjusted ObjectID range
    filtered_list = adobe_json[start_index:end_index]

    # print(filtered_list)

    return filtered_list


def extract_form_numbers_from_json(adobe_json: dict, index: tuple, form_number_patterns):
    text_list = extract_text_from_adobe_json(filter_adobe_json_by_object_id(adobe_json, index[0], index[1]))

    # Convert single pattern to list if needed
    if isinstance(form_number_patterns, str):
        form_number_patterns = [form_number_patterns]

    # Compile all patterns
    compiled_regexes = [re.compile(pattern) for pattern in form_number_patterns]

    form_numbers = []
    for text in text_list:
        text = preprocess_text(text)
        # Try each pattern
        for regex in compiled_regexes:
            matchedlist = regex.findall(text.strip())
            if matchedlist:
                print("MATCHED LIST:", matchedlist)
                form_numbers.extend(matchedlist)

    form_numbers = list(dict.fromkeys(form_numbers))
    return form_numbers


def normalize_pattern(s):
    # Use regular expressions to remove spaces and parentheses
    return re.sub(r"[ ()-]", "", s)


def extractFormNumber(adobe_json, index, regex_pattern):
    # Extract form numbers from each PDF
    form_numbers = extract_form_numbers_from_json(adobe_json, index, regex_pattern)
    print("Form Numbers: ", form_numbers)
    # normalizeFormNumber = [normalize_pattern(i) for i in form_numbers]
    return form_numbers


# Tabular Mode Form Number Processing Code


def xlsx_to_csv_data(xlsx_file, config=None):
    """
    Convert XLSX file content to CSV string format

    Parameters:
    xlsx_file: File-like object containing XLSX data
    config (dict): Configuration dictionary for customization options

    Returns:
    str: CSV formatted string
    """
    csv_data = ""
    workbook = openpyxl.load_workbook(xlsx_file, data_only=True)
    sheet = workbook.active

    try:
        csv_data = ""
        workbook = openpyxl.load_workbook(xlsx_file, data_only=True)
        sheet = workbook.active

        for row in sheet.iter_rows(values_only=True):
            row_data = []
            for cell in row:
                try:
                    if isinstance(cell, str):
                        cell_value = cell.replace("_x000D_", "")
                    else:
                        cell_value = str(cell) if cell is not None else ""
                    row_data.append(cell_value)
                except Exception as e:
                    row_data.append("")
                    print(f"Error processing cell: {e}")
            csv_data += ",".join(row_data) + "\n"

        return csv_data
    except Exception as e:
        raise ValueError(f"Error converting XLSX to CSV: {e}")


def adobe_json_from_zip(zip_path):
    with zipfile.ZipFile(zip_path) as zip_ref:
        target_file = "structuredData.json"
        if target_file not in zip_ref.namelist():
            raise FileNotFoundError(f"Zip does not contain {target_file}")
        with zip_ref.open(target_file) as file:
            adobe_json = json.load(file)
            return adobe_json


import pandas as pd
import zipfile
import os
from io import StringIO
import openpyxl
from io import BytesIO

def read_excel_file(excel_content):
    """
    Read Excel file content and handle potential formatting issues
    """
    try:
        # Read Excel file
        wb = openpyxl.load_workbook(BytesIO(excel_content), data_only=True)
        sheet = wb.active

        # Convert sheet data to list of lists, with improved cell cleaning
        data = []
        for row in sheet.iter_rows(values_only=True):
            row_data = []
            for cell in row:
                try:
                    if isinstance(cell, str):
                        # Clean string values
                        cell_value = cell.replace("_x000D_", "").strip()
                    else:
                        # Convert non-string values
                        cell_value = str(cell) if cell is not None else ""

                    if cell_value:  # Only append non-empty values
                        row_data.append(cell_value)
                except Exception as e:
                    print(f"Error processing cell: {e}")
                    row_data.append("")

            if any(row_data):  # Only add rows that have at least one non-empty value
                data.append(row_data)

        # Find the maximum number of columns
        max_cols = max(len(row) for row in data)

        # Pad rows with None to ensure consistent length
        padded_data = [row + [None] * (max_cols - len(row)) for row in data]

        # Create DataFrame with meaningful column names
        df = pd.DataFrame(padded_data)

        # Clean up the DataFrame
        df = clean_dataframe(df)

        return df

    except Exception as e:
        print(f"Error reading Excel content: {str(e)}")
        return pd.DataFrame()

def clean_dataframe(df):
    """
    Clean and restructure a potentially misaligned DataFrame while preserving column structure
    """
    # Store original column count
    original_column_count = len(df.columns)

    # Drop only empty rows, NOT columns
    df = df.dropna(how='all')

    # Reset the index after dropping rows
    df = df.reset_index(drop=True)

    # Clean up column names if they exist, otherwise use numbered columns
    if all(isinstance(col, str) for col in df.columns):
        df.columns = [str(col).strip() for col in df.columns]
    else:
        df.columns = [f'Column_{i}' for i in range(len(df.columns))]

    # Clean up all string values in the DataFrame
    for column in df.columns:
        df[column] = df[column].apply(lambda x: x.strip() if isinstance(x, str) else x)

    # Fill NaN values with empty strings to preserve structure
    df = df.fillna('')

    # Ensure we preserve the original column structure
    if len(df.columns) < original_column_count:
        print(f"Warning: Column count mismatch. Restoring {original_column_count - len(df.columns)} empty columns")
        for i in range(len(df.columns), original_column_count):
            df[f'Column_{i}'] = ''

    return df

# The rest of your ZipToDF function remains the same

def ZipToDF(zip_path, config=None):
    """
    Read CSV and Excel files from 'tables' folder in a ZIP archive.
    Returns a dictionary with file names as keys and DataFrames as values.
    """
    dataframes = {}

    # Get configuration options with defaults
    tables_folder = config.get("TableConfig", {}).get("TablesFolder", "tables")
    file_extensions = tuple(config.get("TableConfig", {}).get("FileExtensions", (".csv", ".xlsx")))
    encoding = config.get("TableConfig", {}).get("Encoding", "utf-8")
    show_contents = config.get("show_contents", True)

    try:
        with zipfile.ZipFile(zip_path, "r") as zip_ref:
            # Show contents if enabled
            if show_contents:
                print("\nContents of ZIP file:")
                print("-" * 50)
                for file_info in zip_ref.infolist():
                    size = file_info.file_size
                    size_str = f"{size/1024:.1f} KB" if size >= 1024 else f"{size} bytes"
                    print(f"{file_info.filename:<50} {size_str:>10}")
                print("-" * 50 + "\n")

            # Normalize tables_folder path
            normalized_tables_folder = tables_folder.replace("\\", "/").rstrip("/") + "/"

            # Filter target files
            target_files = [
                f for f in zip_ref.namelist()
                if f.lower().endswith(file_extensions) and f.startswith(normalized_tables_folder)
            ]

            if not target_files:
                print(f"No {', '.join(file_extensions)} files found in '{tables_folder}' folder")
                return dataframes

            print(f"Processing files from '{tables_folder}' folder:")
            print("-" * 50)

            # Sort files by their numeric part (e.g., "fileoutpart3.csv" -> 3)
            def extract_number(filename):
                match = re.search(r"part(\d+)", filename)
                return int(match.group(1)) if match else -1

            target_files.sort(key=extract_number)

            for filepath in target_files:
                try:
                    filename = os.path.basename(filepath)
                    print(f"Reading: {filepath}")

                    with zip_ref.open(filepath) as file:
                        content = file.read()

                        if filepath.lower().endswith('.xlsx'):
                            df = read_excel_file(content)
                        else:  # CSV file
                            # Modified CSV handling
                            text_content = content.decode(encoding)
                            import csv
                            from io import StringIO

                            # Use csv.reader with proper quoting settings
                            csv_data = []
                            csv_reader = csv.reader(StringIO(text_content), quoting=csv.QUOTE_ALL)
                            for row in csv_reader:
                                csv_data.append(row)

                            if csv_data:
                                # Find the maximum number of columns
                                max_cols = max(len(row) for row in csv_data)
                                # Create column names based on max columns
                                columns = [f'Column_{i+1}' for i in range(max_cols)]

                                # Create DataFrame and pad shorter rows with None
                                df = pd.DataFrame([row + [None] * (max_cols - len(row)) for row in csv_data],
                                                columns=columns)
                            else:
                                df = pd.DataFrame()

                            # Clean the DataFrame
                            df = clean_dataframe(df)

                        if not df.empty:
                            print(f"DataFrame shape: {df.shape}")
                            print("Columns:", df.columns.tolist())
                            dataframes[filename] = df
                            print(f"Successfully processed: {filename}")
                        else:
                            print(f"Warning: Empty DataFrame for {filename}")

                except Exception as e:
                    print(f"Error processing {filepath}: {str(e)}")
                    continue

    except zipfile.BadZipFile:
        print(f"Error: {zip_path} is not a valid ZIP file")
    except Exception as e:
        print(f"Error processing files: {str(e)}")

    return dataframes


def detect_policy_rows(form_table, config=None):
    """
    Detect rows containing base policy and declaration information based on regex conditions.
    """
    try:
        if config is None or form_table is None:
            raise ValueError("Form table and config must not be None.")

        # Clean column names - remove BOM and strip whitespace
        form_table.columns = [col.replace('\ufeff', '').strip() for col in form_table.columns]
        
        print("\nDEBUG: Column names after cleaning:")
        print(form_table.columns.tolist())

        column_requirements = config.get("TableConfig").get("ColumnRequirements", {})
        base_policy_cond = config.get("TableConfig").get("BasePolicyCondition", {})
        declaration_cond = config.get("TableConfig").get("DeclarationCondition", {})

        # Extract logic
        base_policy_logic = base_policy_cond.get("Logic", "or").lower()
        declaration_logic = declaration_cond.get("Logic", "or").lower()

        def resolve_column_name(col_name):
            try:
                # Get possible alternatives for this column from config
                valid_aliases = [col_name] + config.get("TableConfig", {}).get("ColumnRequirements", {}).get(col_name, [])
                
                # Convert form_table columns to a set of cleaned column names
                form_table_cols = {col.upper().strip(): col for col in form_table.columns}
                
                print(f"\nDEBUG: Resolving column {col_name}")
                print(f"Valid aliases: {valid_aliases}")
                print(f"Available columns: {form_table_cols}")
                
                # Try to find a match
                for alias in valid_aliases:
                    alias_clean = alias.upper().strip()
                    if alias_clean in form_table_cols:
                        resolved = form_table_cols[alias_clean]
                        print(f"Found match: {resolved}")
                        return resolved
                        
                print(f"No match found for {col_name}")
                return None
            except Exception as e:
                print(f"Error resolving column name '{col_name}': {e}")
                return None

        def apply_conditions(conditions_dict, logic_op, condition_type):
            masks = []
            print(f"\nApplying {condition_type} conditions with logic='{logic_op.upper()}':")
            
            for column, pattern in conditions_dict.items():
                if column.lower() == 'logic':
                    continue  # Skip the logic key
                    
                try:
                    # Resolve the column name using the config
                    resolved_col = resolve_column_name(column)
                    
                    if resolved_col and resolved_col in form_table.columns:
                        # Convert column to string, strip whitespace, then apply regex
                        col_series_str = form_table[resolved_col].astype(str).str.strip()
                        
                        # Check for valid regex pattern
                        try:
                            mask = col_series_str.str.contains(pattern, regex=True, na=False, case=False)
                        except Exception as regex_err:
                            print(f"Invalid regex pattern '{pattern}' for column '{resolved_col}': {regex_err}")
                            continue

                        # Log info
                        matched_count = mask.sum()
                        total_rows = len(form_table)
                        print(f" - Condition on column '{resolved_col}' (resolved from '{column}') " 
                              f"with pattern '{pattern}': {matched_count} out of {total_rows} rows matched.")

                        masks.append(mask)
                    else:
                        print(f" - Skipped '{column}' as no matching column found in DataFrame.")
                except Exception as e:
                    print(f"Error applying condition for column '{column}': {e}")
                    masks.append(pd.Series(False, index=form_table.index))

            if not masks:
                print(f" - No applicable columns found for {condition_type} conditions.")
                return pd.Series(False, index=form_table.index)

            combined_mask = pd.concat(masks, axis=1)
            if logic_op == "and":
                final_mask = combined_mask.all(axis=1)
            else:  # "or"
                final_mask = combined_mask.any(axis=1)

            final_matched_count = final_mask.sum()
            print(f" - Combined result ({logic_op.upper()} logic): {final_matched_count} out of {len(form_table)} rows matched {condition_type} conditions.\n")
            return final_mask

        # Remove the 'logic' key from conditions before applying
        base_conditions = {k: v for k, v in base_policy_cond.items() if k.lower() != 'logic'}
        dec_conditions = {k: v for k, v in declaration_cond.items() if k.lower() != 'logic'}

        base_mask = apply_conditions(base_conditions, base_policy_logic, "base policy")
        base_policy_rows = form_table[base_mask]

        dec_mask = apply_conditions(dec_conditions, declaration_logic, "declaration")
        declaration_rows = form_table[dec_mask]

        return base_policy_rows, declaration_rows

    except Exception as e:
        print(f"Error in detect_policy_rows function: {e}")
        return None, None


def split_dict_to_ordering_dict(base_rows, dec_rows, split_dict, config=None):
    """
    Create ordering dictionary by matching form numbers from base and declaration rows
    to entries in the split dictionary.
    """
    # Initialize result dictionary
    ordering_dict = {}

    # Debug print the split dictionary contents
    print("\nDEBUG: Split Dictionary Contents:")
    print("-" * 50)
    for key, value in split_dict.items():
        print(f"Key: {key}")
        print(f"Text: {value['text']}")
        print(f"Normalized: {normalize_text_for_matching(value['text'])}")
        print("-" * 30)

    def resolve_column_name(col_name, df):
        df_cols = {c.strip().upper(): c for c in df.columns}
        valid_aliases = [col_name] + config.get("TableConfig", {}).get("ColumnRequirements", {}).get(col_name, [])
        for alias in valid_aliases:
            alias_clean = alias.strip().upper()
            if alias_clean in df_cols:
                return df_cols[alias_clean]
        return col_name

    def find_index_in_split_dict(form_number):
        """Helper function to find index for a form number in split dict"""
        try:
            # Print the input form number and its normalized version
            print(f"\nDEBUG: Trying to match form number: {form_number}")
            normalized_form = normalize_text_for_matching(form_number)
            print(f"DEBUG: Normalized form number: {normalized_form}")
            
            # Create and print the mapping of normalized split dict text to their indices
            split_dict_mapping = {
                normalize_text_for_matching(value["text"]): int(key.split("_")[-1])
                for key, value in split_dict.items()
            }
            
            print("\nDEBUG: Available normalized split dict entries:")
            for text, idx in split_dict_mapping.items():
                print(f"Index {idx}: {text}")
            
            if normalized_form in split_dict_mapping:
                matched_idx = split_dict_mapping[normalized_form]
                print(f"DEBUG: Found match at index {matched_idx}")
                return matched_idx
            else:
                print("DEBUG: No match found in split dictionary")
            return None
        except Exception as e:
            print(f"Error in find_index_in_split_dict: {str(e)}")
            return None

    # Process base policy rows
    try:
        if not base_rows.empty:
            print("\nDEBUG: Processing Base Policy Rows")
            print("-" * 50)
            print(f"Base rows columns: {base_rows.columns.tolist()}")
            
            form_number_col = resolve_column_name("FORM NUMBER", base_rows)
            print(f"Resolved form number column: {form_number_col}")
            
            if form_number_col in base_rows.columns:
                print("\nDEBUG: Base Policy Form Numbers:")
                base_form_numbers = base_rows[form_number_col].dropna().unique()
                for form_number in base_form_numbers:
                    print(f"\nProcessing base form number: {form_number}")
                    print(f"Normalized version: {normalize_text_for_matching(form_number)}")
                    base_index = find_index_in_split_dict(form_number)
                    if base_index is not None:
                        print(f"Found base policy '{form_number}' at index {base_index}")
                        if "OrderingKey" not in ordering_dict:
                            ordering_dict["OrderingKey"] = {}
                        ordering_dict["OrderingKey"]["Base"] = base_index
                    else:
                        print(f"Warning: Base policy form number '{form_number}' not found in split dictionary")
            else:
                print(f"Warning: Could not find form number column in base policy rows. Available columns: {base_rows.columns.tolist()}")
    except Exception as e:
        print(f"Error processing base policy rows: {str(e)}")

    # Process declaration rows
    try:
        if not dec_rows.empty:
            print("\nDEBUG: Processing Declaration Rows")
            print("-" * 50)
            print(f"Declaration rows columns: {dec_rows.columns.tolist()}")
            
            form_number_col = resolve_column_name("FORM NUMBER", dec_rows)
            print(f"Resolved form number column: {form_number_col}")
            
            if form_number_col in dec_rows.columns:
                print("\nDEBUG: Declaration Form Numbers:")
                dec_form_numbers = dec_rows[form_number_col].dropna().unique()
                for form_number in dec_form_numbers:
                    print(f"\nProcessing declaration form number: {form_number}")
                    print(f"Normalized version: {normalize_text_for_matching(form_number)}")
                    dec_index = find_index_in_split_dict(form_number)
                    if dec_index is not None:
                        print(f"Found declaration '{form_number}' at index {dec_index}")
                        if "OrderingKey" not in ordering_dict:
                            ordering_dict["OrderingKey"] = {}
                        ordering_dict["OrderingKey"].setdefault("Declaration", []).append(dec_index)
                    else:
                        print(f"Warning: Declaration form number '{form_number}' not found in split dictionary")
            else:
                print(f"Warning: Could not find form number column in declaration rows. Available columns: {dec_rows.columns.tolist()}")
    except Exception as e:
        print(f"Error processing declaration rows: {str(e)}")

    print("\nDEBUG: Final Ordering Dictionary:")
    print(ordering_dict)
    return ordering_dict

def normalize_text_for_matching(text):
    """
    Normalize text by removing spaces, parentheses, and converting to lowercase
    for consistent matching.
    """
    """Remove spaces, parentheses, and other special characters."""
    original = text
    normalized = re.sub(r"[ ()-/]", "", text).lower()
    print(f"DEBUG normalize_pattern: {original} -> {normalized}")
    return normalized



def extract_form_numbers(form_table, config=None):
    """
    Extract a list of form numbers from the form table based on the form_number_pattern.

    Parameters:
    form_table (pd.DataFrame): DataFrame containing form information
    config (dict): Configuration dictionary containing the form_number_pattern

    Returns:
    list: List of matched form numbers
    """
    try:
        # Ensure config is provided, else return an empty list
        if config is None:
            print("Warning: No configuration provided.")
            return []

        form_number_pattern = config.get("FormNumberPattern", None)

        # Check if form_number_pattern exists and 'FORM NUMBER' column is in form_table
        if not form_number_pattern:
            print("Warning: No form_number_pattern found in configuration.")
            return []
        if "FORM NUMBER" not in form_table.columns:
            print("Warning: 'FORM NUMBER' column not found in form table.")
            return []

        # Attempt to extract form numbers
        form_numbers = form_table["FORM NUMBER"].astype(str).str.strip().tolist()

        # Use regex to filter valid form numbers
        try:
            pattern = re.compile(form_number_pattern)
            matched_form_numbers = [fn for fn in form_numbers if pattern.match(fn)]
        except re.error as e:
            print(f"Error compiling regex pattern '{form_number_pattern}': {str(e)}")
            return []

        return matched_form_numbers

    except Exception as e:
        print(f"Error in extract_form_numbers: {str(e)}")
        return []


def preprocess_string(s):
    # Remove or replace unwanted characters
    return re.sub(r"[\x00-\x1F\x7F\xAD]", "-", s).strip()


def detect_form_number_table(dataframes_dict, config):
    """
    Detect and extract relevant rows containing form numbers from multiple DataFrames.
    Uses regex patterns and column positions for matching split tables.
    """
    try:
        # Ensure config contains the necessary patterns
        if not config or "FormNumberPattern" not in config:
            print("Error: Configuration does not contain 'FormNumberPattern'")
            return pd.DataFrame()

        # Get column requirements from config
        column_requirements = config.get("TableConfig", {}).get("ColumnRequirements", {})
        form_number_aliases = column_requirements.get("FORM NUMBER", ["FORM NUMBER"])
        name_aliases = column_requirements.get("NAME", ["NAME"])
        edition_date_aliases = column_requirements.get("EDITION DATE", ["EDITION DATE"])

        # Get patterns from config and handle both string and list inputs
        form_number_patterns = config["FormNumberPattern"]
        if isinstance(form_number_patterns, str):
            form_number_patterns = [form_number_patterns]

        def find_header_row(df):
            """Find the row that contains our column headers by looking for matching content"""
            aliases = {
                'FORM NUMBER': form_number_aliases,
                'NAME': name_aliases,
                'EDITION DATE': edition_date_aliases
            }

            for idx in range(len(df)):
                row = df.iloc[idx]
                column_mapping = {}
                matches_found = 0

                # Check each cell in the row
                for col in df.columns:
                    cell_value = str(row[col]).strip()
                    if cell_value.startswith('\ufeff'):  # Remove BOM if present
                        cell_value = cell_value[1:]

                    # Check against each type of column alias
                    for col_type, alias_list in aliases.items():
                        if any(alias.strip().lower() == cell_value.lower() for alias in alias_list):
                            column_mapping[col_type] = col
                            matches_found += 1
                            break

                print("Column mapping: ", column_mapping)

                # We need at least FORM NUMBER and NAME columns
                if 'FORM NUMBER' in column_mapping and 'NAME' in column_mapping:
                    return idx, column_mapping

            return None, None

        def match_form_number(value):
            """Check if value matches any of the form number patterns"""
            value = str(value).strip()
            if value.startswith('\ufeff'):
                value = value[1:]
            return any(re.match(pattern, value) for pattern in form_number_patterns)

        # Process each DataFrame in the dictionary
        merged_df = pd.DataFrame()
        main_table_found = False
        main_table_info = None

        # Sort files by their numeric part
        sorted_files = sorted(dataframes_dict.keys(),
                            key=lambda x: int(re.search(r"part(\d+)", x).group(1)))

        for filename in sorted_files:
            print(f"\nProcessing file: {filename}")
            df = dataframes_dict[filename].copy()

            if not main_table_found:
                # Original initial table detection logic
                header_idx, column_mapping = find_header_row(df)

                if header_idx is None:
                    print(f"Skipping {filename}: No valid header row found")
                    continue

                print("\nDEBUG - Header Information:")
                print("Original df columns:", df.columns.tolist())
                print("Header index:", header_idx)
                print("Column mapping:", column_mapping)

                form_col = column_mapping['FORM NUMBER']
                print("form_col value:", form_col)

                print("Header row full data:")
                print(df.iloc[header_idx])

                # Extract the table with proper headers
                header_row = df.iloc[header_idx]
                data_df = df.iloc[header_idx + 1:].reset_index(drop=True)

                print("\nBefore column renaming:")
                print("data_df columns:", data_df.columns.tolist())

                # Check form numbers using original column names
                form_number_matches = data_df[form_col].astype(str).apply(match_form_number)

                if not form_number_matches.any():
                    print(f"Skipping {filename}: No form numbers matching the patterns found")
                    continue

                # Create new_columns dictionary based on actual header values
                new_columns = {}
                for col in df.columns:
                    new_columns[col] = header_row[col].strip()

                # Rename the columns
                data_df = data_df.rename(columns=new_columns)

                # Store the table information for future matches
                main_table_info = {
                    'column_mapping': column_mapping,
                    'header_row': header_row,
                    'original_columns': df.columns.tolist(),
                    'new_columns': new_columns
                }

                print(f"Main table established from {filename}")
                main_table_found = True

                # Create merged_df with the matching rows and new column names
                merged_df = data_df[form_number_matches].reset_index(drop=True)

            else:
                print(f"\nAttempting to merge table from {filename}")
                # Find header row in subsequent file
                header_idx, column_mapping = find_header_row(df)

                if header_idx is not None:
                    print("Found header row in subsequent file")
                    if 'FORM NUMBER' not in column_mapping:
                        print(f"Error: Could not find FORM NUMBER column in subsequent file {filename}")
                        return merged_df
                    form_col = column_mapping['FORM NUMBER']
                    data_df = df.iloc[header_idx + 1:].reset_index(drop=True)
                else:
                    print("No header row found - treating as direct continuation")
                    # Try to find form number column based on main table mapping
                    if 'FORM NUMBER' not in main_table_info['column_mapping']:
                        print(f"Error: Could not find FORM NUMBER column mapping for continuation in {filename}")
                        return merged_df
                    form_col = main_table_info['column_mapping']['FORM NUMBER']
                    data_df = df.copy()

                print(f"Checking form numbers in column: {form_col}")
                print("Sample values:")
                print(data_df[form_col].head())

                # Check form numbers using the resolved column name
                form_number_matches = data_df[form_col].astype(str).apply(match_form_number)

                if not form_number_matches.any():
                    print(f"No matching form numbers found in {filename}")
                    return merged_df

                print(f"Found {form_number_matches.sum()} matching form numbers")

                # Apply the column renaming from main table
                data_df = data_df.rename(columns=main_table_info['new_columns'])

                # If we found a header row, include it in the merge
                if header_idx is not None:
                    header_df = pd.DataFrame([df.iloc[header_idx]], columns=merged_df.columns)
                    merged_df = pd.concat([merged_df, header_df, data_df[form_number_matches]], ignore_index=True)
                else:
                    # Direct concatenation for continuation data
                    if len(data_df.columns) != len(merged_df.columns):
                        print(f"Column count mismatch in {filename}. Expected: {len(merged_df.columns)}, Got: {len(data_df.columns)}")
                        return merged_df
                    merged_df = pd.concat([merged_df, data_df[form_number_matches]], ignore_index=True)

                print(f"Successfully merged data from {filename}")
                print(f"Current merged DataFrame shape: {merged_df.shape}")

        if not main_table_found:
            print("Warning: No relevant tables found in any DataFrame")
            return pd.DataFrame()

        print("\nProcessing completed")
        print(f"Final merged DataFrame shape: {merged_df.shape}")
        return merged_df

    except Exception as e:
        print(f"Error in detect_form_number_table: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return pd.DataFrame()


def process_ordering(zip_path, config=None, split_dict=None):
    """
    Process form ordering from ZIP files containing CSV/Excel files.

    Parameters:
    zip_path (str): Path to the ZIP file
    config (dict): Configuration dictionary
    split_dict (dict, optional): Dictionary containing split information

    Returns:
    tuple: (result, status) where result is either a list of form numbers or a dictionary
    """
    try:
        # Ensure config is not None
        if config is None:
            print("Warning: Configuration is None, using default empty config.")
            config = {}

        # Handle text mode
        if config.get("Mode", "").lower().strip() == "text":
            if "FormNumberPattern" in config:
                try:
                    adobe_json = adobe_json_from_zip(zip_path=zip_path)
                    patterns = config.get("FormNumberPattern")
                    if isinstance(patterns, str):
                        patterns = [patterns]
                    form_numbers = extractFormNumber(adobe_json=adobe_json, index=(None, None), regex_pattern=patterns)
                    return (form_numbers), ("Success")
                except FileNotFoundError as e:
                    return (), (str(e))
                except Exception as e:
                    return (), (f"Error in text mode processing: {str(e)}")

        if not os.path.exists(zip_path):
            return ({}), (f"Error: The ZIP file at {zip_path} does not exist.")

        # Get column aliases from config
        column_requirements = config.get("TableConfig", {}).get("ColumnRequirements", {})
        form_number_aliases = column_requirements.get("FORM NUMBER", ["FORM NUMBER"])

        # Attempt to read the ZIP file and convert to DataFrames
        try:
            print("\nDEBUG: Reading ZIP file")
            dfs = ZipToDF(zip_path, config)
            print(f"Found {len(dfs)} DataFrames in ZIP")
        except Exception as e:
            return ({}), (f"Error processing ZIP file '{zip_path}': {str(e)}")

        # Attempt to detect form number table
        print("\nDEBUG: Detecting form number table")
        form_number_table = detect_form_number_table(dfs, config)

        if form_number_table is not None and not form_number_table.empty:
            print(f"\nDEBUG: Found form number table with shape {form_number_table.shape}")
            # First extract all form numbers using existing logic
            form_numbers = set()
            patterns = config.get("FormNumberPattern", "")
            if isinstance(patterns, str):
                patterns = [patterns]

            try:
                print("\nDEBUG: Starting form number extraction")
                regexes = [re.compile(pattern) for pattern in patterns]
                date_config = config.get("DateSubConfig", None)
                
                print(f"\nDate config present: {date_config is not None}")
                if date_config:
                    print("Date config:", json.dumps(date_config, indent=2))

                # Create a dictionary to store form numbers with their standardized versions
                standardized_form_numbers = {}

                for column in form_number_table.columns:
                    print(f"\nProcessing column: {column}")
                    column_data = form_number_table[column].astype(str).str.strip()
                    column_data = column_data.str.replace("\n", " ").str.replace("\r", "")
                    column_data = column_data.apply(lambda x: preprocess_string(str(x)))
                    column_data = column_data.str.replace(r"\s+", " ", regex=True)

                    for regex in regexes:
                        matches = column_data.str.match(regex, na=False)
                        matching_values = column_data[matches].unique()

                        for val in matching_values:
                            if val and not val.isspace():
                                standardized_val = val
                                if date_config:
                                    try:
                                        print(f"\nDEBUG Processing form number: {val}")
                                        date_col, is_embedded = identify_date_column_strategy(form_number_table, config)
                                        print(f"Date strategy - is_embedded: {is_embedded}, date_col: {date_col}")

                                        if is_embedded and column in form_number_aliases:
                                            print("Processing embedded date")
                                            clean_form_num, date_str, standardized_date = extract_date_from_form_number(val, date_config)
                                            print(f"Clean form: {clean_form_num}")
                                            print(f"Date string: {date_str}")
                                            print(f"Standardized date: {standardized_date}")
                                            if standardized_date:
                                                standardized_val = clean_form_num + " (" + standardized_date + ")"
                                                print(f"Standardized form number: {standardized_val}")
                                        elif date_col and not is_embedded:
                                            print("Processing separate date column")
                                            matching_rows = form_number_table[
                                                form_number_table[column].astype(str).str.strip() == val
                                            ]
                                            if not matching_rows.empty:
                                                date_val = matching_rows[date_col].iloc[0]
                                                if date_val and str(date_val).strip():
                                                    original_date, standardized_date = extract_and_standardize_date(
                                                        str(date_val),
                                                        date_config["table"]["pattern"],
                                                        date_config["table"]["format"],
                                                        date_config["footer"]["format"]
                                                    )
                                                    if standardized_date:
                                                        standardized_val = combine_form_number_with_date(
                                                            val.strip(),
                                                            f"({standardized_date})",
                                                            date_config
                                                        )
                                                        print(f"Standardized form number with separate date: {standardized_val}")
                                    except Exception as e:
                                        print(f"Warning: Error processing date for {val}: {e}")
                                
                                standardized_form_numbers[val] = standardized_val
                                form_numbers.add(standardized_val)
                                print(f"Added form number: {standardized_val}")

                # If split_dict is provided, use the standardized form numbers for matching
                if split_dict:
                    try:
                        print("\nDEBUG: Creating processed table with standardized form numbers")
                        processed_table = form_number_table.copy()
                        print(processed_table)
                        
                        # Update the form numbers with their standardized versions
                        for idx, row in processed_table.iterrows():
                            for col in form_number_aliases:
                                if col in processed_table.columns:
                                    val = str(row[col]).strip()
                                    if val in standardized_form_numbers:
                                        print(f"Updating form number: {val} -> {standardized_form_numbers[val]}")
                                        processed_table.at[idx, col] = standardized_form_numbers[val]

                        base_policy_rows, declaration_rows = detect_policy_rows(processed_table, config)
                        ordering_dict = split_dict_to_ordering_dict(base_policy_rows, declaration_rows, split_dict, config)
                        return ordering_dict, "Success"
                    except Exception as e:
                        print(f"Error creating ordering dict: {str(e)}")
                        return {}, f"Error during ordering dict creation: {str(e)}"
                else:
                    # Return sorted list of form numbers if no split_dict
                    form_numbers = sorted(list(form_numbers))
                    print("\nFinal Form Numbers:")
                    for fn in form_numbers:
                        print(f"  - {fn}")
                    return (form_numbers), ("Success")

            except Exception as e:
                print(f"Error in form number processing: {str(e)}")
                return ({} if split_dict else []), f"Error extracting form numbers from table: {str(e)}"

        return ({} if split_dict else []), "No form number table found with specified criteria"

    except Exception as e:
        print(f"Error in process_ordering: {str(e)}")
        return ({} if split_dict else []), f"Error in process_ordering: {str(e)}"

# =============================================================================
#                                  MAIN
# =============================================================================
if __name__ == "__main__":

    # config, split_dict = {}

    # config = {
    #     "CarrierName": "Kemper",
    #     "Mode": "Text",
    #     "FormNumberPattern": r"\b(?:EZ-\d|\w{2,3}\d{4}[A-Z]?)\s?\(\d{2}/\d{2}\)",
    # }

    config = {
        "CarrierName": "pure",
        "Mode": "Table",
        "TableConfig": {
            "TablesFolder": "tables",
            "FileExtensions": [".csv", ".xlsx"],
            "Encoding": "utf-8",
            "ColumnRequirements": {"FORM NUMBER": ["FORM NO", "FORM NUMBER", "FORM ID"], "NAME": ["FORMS & ENDORSEMENTS", "FORM NAME", "DESCRIPTION"], "EDITION DATE": ["ED DATE", "EDITION DATE", "DATE"]},
            "BasePolicyCondition": {"logic": "or", "NAME": "policy$", "FORM NUMBER": "PCF"},
            "DeclarationCondition": {"logic": "or", "NAME": "declaration", "FORM NUMBER": "DEC"},
        },
        "FormNumberPattern": "^[A-Z]+-(?:\\d{3}(?:-[A-Z]+)?|[A-Z]+-(?:[A-Z]+)-\\d{3}|[A-Z]+-[A-Z]+|[A-Z]+)(?:\\s*\\(?\\d{1,2}(?:\\/\\d{2,4}|\\-\\d{2,4}|\\.\\d{2,4})\\)?)?$",
    }

    split_dict = {
        "123_0": {"text": "PHVH-DEC-KS-001 (08/2045)", "split": (0, 313), "page_num": 3},
        "123_1": {"text": "PURE-DSC-GEN-001 (08/2015)", "split": (314, 333), "page_num": 4},
        "123_2": {"text": "PURE-001-OR (03/2011)", "split": (334, 343), "page_num": 5},
        "123_3": {"text": "PHVH-PCF-GEN-001 (02/2020)", "split": (344, 1026), "page_num": 33},
        "123_4": {"text": "PHVH-END-OR-001 (12/2022)", "split": (1027, 1223), "page_num": 39},
        "123_5": {"text": "PHVH-END-GEN-022 (02/2020)", "split": (1224, 1359), "page_num": 47},
        "123_6": {"text": "PHVH-043-GEN (03/2008)", "split": (1360, 1361), "page_num": 48},
        "123_7": {"text": "PURE-002 (08/2008)", "split": (1361, 1366), "page_num": 49},
        "123_8": {"text": "PURE-003 (03/2009)", "split": (1367, 1378), "page_num": 50},
        "123_9": {"text": "PURE-038-GEN (03/2014)", "split": (1379, 1465), "page_num": 54},
    }

    # # 1. With split_dict to get the ordering dictionary
    # print("=== Processing with split_dict to get Ordering Dictionary ===")
    # ordering_result = process_ordering(r"E:\Projects\ML\annotation-script\data\pure\adobe-api-Scenario 5 - Prompt-932568f2-d7f3-407d-81d0-68615d49d78d.zip", config=config, split_dict=split_dict)
    # print("\nFinal Ordering Dictionary:")
    # print("Ordering Result:", ordering_result)

    # 2. Without split_dict to get the list of form numbers
    print("\n=== Processing without split_dict to get Form Numbers ===")
    form_numbers_result = process_ordering(zip_path="/content/adobe-api-Scenario 5 - Prompt-932568f2-d7f3-407d-81d0-68615d49d78d.zip", config=config, split_dict=split_dict)
    print("Form Numbers:", form_numbers_result)
    # form_numbers_result = process_ordering(zip_path=r"E:\Projects\ML\annotation-script\data\DEC Page example.zip", config=config)